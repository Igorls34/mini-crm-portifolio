from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.models import User
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.http import HttpResponse, JsonResponse
import csv
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.units import inch
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from .models import Lead, ActivityLog
from .forms import LeadForm
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.models import User
import json


# Autenticação
def login_view(request):
    if request.method == "POST":
        username = request.POST["username"]
        password = request.POST["password"]
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            next_url = request.POST.get('next') or 'dashboard'
            return redirect(next_url)
        else:
            messages.error(request, "Usuário ou senha incorretos. Tente novamente.")
    return render(request, "login.html")


def logout_view(request):
    logout(request)
    return redirect("login")


def register_view(request):
    print(f"[DEBUG] Register view chamada - Método: {request.method}")
    if request.method == "POST":
        form = UserCreationForm(request.POST)
        print(f"[DEBUG] Formulário submetido - Válido: {form.is_valid()}")
        if form.is_valid():
            user = form.save()
            login(request, user)
            print(f"[DEBUG] Usuário {user.username} registrado e logado")
            messages.success(
                request, f"Bem-vindo, {user.username}! Conta criada com sucesso."
            )
            return redirect("dashboard")
    else:
        form = UserCreationForm()
    return render(request, "register.html", {"form": form})


@login_required
def dashboard(request):
    # Cards existentes
    total_leads = Lead.objects.count()
    novos = Lead.objects.filter(status="novo").count()
    progresso = Lead.objects.filter(status="progresso").count()
    convertidos = Lead.objects.filter(status="convertido").count()

    # Novos cards
    from django.db.models import Avg, Sum
    ticket_medio = Lead.objects.filter(valor_curso__isnull=False).aggregate(avg=Avg('valor_curso'))['avg'] or 0
    faturamento_total = Lead.objects.filter(status="convertido", valor_curso__isnull=False).aggregate(total=Sum('valor_curso'))['total'] or 0

    # Últimos leads
    ultimos_leads = Lead.objects.order_by("-data_criacao")[:5]

    # Dados para gráfico de status (pipeline donut)
    status_counts = (
        Lead.objects.values("status").annotate(count=Count("status")).order_by("status")
    )
    status_labels = [dict(Lead.STATUS_CHOICES)[s["status"]] for s in status_counts]
    status_data = [s["count"] for s in status_counts]

    # Dados para gráfico de origem
    origem_counts = (
        Lead.objects.values("origem").annotate(count=Count("origem")).order_by("-count")
    )
    origem_labels = [dict(Lead.ORIGEM_CHOICES)[o["origem"]] for o in origem_counts]
    origem_data = [o["count"] for o in origem_counts]

    # Ranking de atendentes
    atendente_ranking = (
        Lead.objects.filter(atendente__isnull=False)
        .values("atendente__username", "atendente__first_name", "atendente__last_name")
        .annotate(convertidos=Count("id", filter=Q(status="convertido")))
        .order_by("-convertidos")[:5]
    )

    context = {
        "total_leads": total_leads,
        "novos": novos,
        "progresso": progresso,
        "convertidos": convertidos,
        "ticket_medio": ticket_medio,
        "faturamento_total": faturamento_total,
        "ultimos_leads": ultimos_leads,
        "status_labels": json.dumps(status_labels),
        "status_data": json.dumps(status_data),
        "origem_labels": json.dumps(origem_labels),
        "origem_data": json.dumps(origem_data),
        "atendente_ranking": atendente_ranking,
    }
    return render(request, "dashboard.html", context)


@login_required
def lead_pipeline(request):
    # Filtrar por atendente se for ATENDENTE
    leads = Lead.objects.all()
    if request.user.groups.filter(name='ATENDENTE').exists():
        leads = leads.filter(atendente=request.user)

    # Organizar leads por status para o Kanban
    kanban_data = {}
    for status_key, status_label in Lead.STATUS_CHOICES:
        kanban_data[status_key] = {
            'label': status_label,
            'leads': leads.filter(status=status_key).order_by('-data_criacao')
        }

    context = {
        'kanban_data': kanban_data,
    }
    return render(request, "leads/pipeline.html", context)


@login_required
def activity_logs(request):
    # Filtrar logs por permissões
    logs = ActivityLog.objects.select_related('lead', 'user')

    if request.user.groups.filter(name='ATENDENTE').exists():
        # ATENDENTE vê apenas logs de leads que são dele
        logs = logs.filter(lead__atendente=request.user)

    # Filtros
    lead_filter = request.GET.get('lead', '')
    user_filter = request.GET.get('user', '')
    action_filter = request.GET.get('action', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if lead_filter:
        logs = logs.filter(lead__nome__icontains=lead_filter)

    if user_filter:
        logs = logs.filter(user__username__icontains=user_filter)

    if action_filter:
        logs = logs.filter(action=action_filter)

    if date_from:
        logs = logs.filter(timestamp__date__gte=date_from)

    if date_to:
        logs = logs.filter(timestamp__date__lte=date_to)

    # Paginação
    paginator = Paginator(logs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'lead_filter': lead_filter,
        'user_filter': user_filter,
        'action_filter': action_filter,
        'date_from': date_from,
        'date_to': date_to,
        'action_choices': ActivityLog.ACTION_CHOICES,
    }
    return render(request, "leads/activity_logs.html", context)


@login_required
def update_lead_status(request, pk):
    if request.method == 'POST':
        try:
            lead = get_object_or_404(Lead, pk=pk)
            # Tentar obter dados JSON primeiro, depois dados de formulário
            if request.content_type == 'application/json':
                import json
                data = json.loads(request.body)
                new_status = data.get('status')
            else:
                new_status = request.POST.get('status')

            old_status = lead.status

            # Verificar permissões
            if request.user.groups.filter(name='ATENDENTE').exists() and lead.atendente != request.user:
                return JsonResponse({'success': False, 'error': 'Permissão negada'})

            if new_status in dict(Lead.STATUS_CHOICES) and old_status != new_status:
                lead.status = new_status
                lead.save()

                # Criar log de atividade
                ActivityLog.objects.create(
                    lead=lead,
                    user=request.user,
                    action='status_changed',
                    old_value=old_status,
                    new_value=new_status,
                    description=f'Status alterado via pipeline de "{dict(Lead.STATUS_CHOICES)[old_status]}" para "{dict(Lead.STATUS_CHOICES)[new_status]}"'
                )

                return JsonResponse({'success': True})
            else:
                return JsonResponse({'success': False, 'error': 'Status inválido'})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Método não permitido'})


@login_required
def lead_list(request):
    query = request.GET.get("q", "")
    status_filter = request.GET.get("status", "")
    origem_filter = request.GET.get("origem", "")
    atendente_filter = request.GET.get("atendente", "")
    prioridade_filter = request.GET.get("prioridade", "")
    prob_min = request.GET.get("prob_min", "")
    prob_max = request.GET.get("prob_max", "")
    data_inicial = request.GET.get("data_inicial", "")
    data_final = request.GET.get("data_final", "")
    order_by = request.GET.get("order", "-data_criacao")

    leads = Lead.objects.all()

    # Filtrar por atendente se for ATENDENTE
    if request.user.groups.filter(name='ATENDENTE').exists():
        leads = leads.filter(atendente=request.user)

    if query:
        leads = leads.filter(Q(nome__icontains=query) | Q(email__icontains=query) | Q(curso_interesse__icontains=query))

    if status_filter:
        leads = leads.filter(status=status_filter)

    if origem_filter:
        leads = leads.filter(origem=origem_filter)

    if atendente_filter:
        leads = leads.filter(atendente_id=atendente_filter)

    if prioridade_filter:
        leads = leads.filter(prioridade=prioridade_filter)

    if prob_min:
        leads = leads.filter(probabilidade_fechamento__gte=prob_min)

    if prob_max:
        leads = leads.filter(probabilidade_fechamento__lte=prob_max)

    if data_inicial:
        leads = leads.filter(data_criacao__date__gte=data_inicial)

    if data_final:
        leads = leads.filter(data_criacao__date__lte=data_final)

    leads = leads.order_by(order_by)

    paginator = Paginator(leads, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "query": query,
        "status_filter": status_filter,
        "origem_filter": origem_filter,
        "atendente_filter": atendente_filter,
        "prioridade_filter": prioridade_filter,
        "prob_min": prob_min,
        "prob_max": prob_max,
        "data_inicial": data_inicial,
        "data_final": data_final,
        "order_by": order_by,
        "users": User.objects.all(),
        "status_choices": Lead.STATUS_CHOICES,
        "origem_choices": Lead.ORIGEM_CHOICES,
        "prioridade_choices": Lead.PRIORIDADE_CHOICES,
    }
    return render(request, "leads/list.html", context)

    leads = leads.order_by(order_by)

    paginator = Paginator(leads, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "query": query,
        "status_filter": status_filter,
        "origem_filter": origem_filter,
        "atendente_filter": atendente_filter,
        "prioridade_filter": prioridade_filter,
        "prob_min": prob_min,
        "prob_max": prob_max,
        "data_inicial": data_inicial,
        "data_final": data_final,
        "order_by": order_by,
        "users": User.objects.all(),
        "status_choices": Lead.STATUS_CHOICES,
        "origem_choices": Lead.ORIGEM_CHOICES,
        "prioridade_choices": Lead.PRIORIDADE_CHOICES,
    }
    return render(request, "leads/list.html", context)


@login_required
def export_leads_xlsx(request):
    # Aplicar os mesmos filtros da listagem
    leads = Lead.objects.all()

    # Filtrar por atendente se for ATENDENTE
    if request.user.groups.filter(name='ATENDENTE').exists():
        leads = leads.filter(atendente=request.user)

    # Aplicar filtros da requisição
    query = request.GET.get("q", "")
    status_filter = request.GET.get("status", "")
    origem_filter = request.GET.get("origem", "")
    atendente_filter = request.GET.get("atendente", "")
    prioridade_filter = request.GET.get("prioridade", "")
    prob_min = request.GET.get("prob_min", "")
    prob_max = request.GET.get("prob_max", "")
    data_inicial = request.GET.get("data_inicial", "")
    data_final = request.GET.get("data_final", "")

    if query:
        leads = leads.filter(Q(nome__icontains=query) | Q(email__icontains=query) | Q(curso_interesse__icontains=query))

    if status_filter:
        leads = leads.filter(status=status_filter)

    if origem_filter:
        leads = leads.filter(origem=origem_filter)

    if atendente_filter:
        leads = leads.filter(atendente_id=atendente_filter)

    if prioridade_filter:
        leads = leads.filter(prioridade=prioridade_filter)

    if prob_min:
        leads = leads.filter(probabilidade_fechamento__gte=prob_min)

    if prob_max:
        leads = leads.filter(probabilidade_fechamento__lte=prob_max)

    if data_inicial:
        leads = leads.filter(data_criacao__date__gte=data_inicial)

    if data_final:
        leads = leads.filter(data_criacao__date__lte=data_final)

    # Criar workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    data_font = Font(size=10)
    data_alignment = Alignment(horizontal="left", vertical="center")

    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Cabeçalhos
    headers = [
        'Nome', 'Telefone', 'Email', 'Curso', 'Origem', 'Prioridade',
        'Probabilidade', 'Valor Curso', 'Atendente', 'Status', 'Data Criação', 'Observação Interna'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Dados
    for row_num, lead in enumerate(leads, 2):
        data = [
            lead.nome,
            lead.telefone,
            lead.email or '',
            lead.curso_interesse,
            lead.get_origem_display(),
            lead.get_prioridade_display(),
            f"{lead.probabilidade_fechamento}%",
            lead.valor_curso if lead.valor_curso else '',
            lead.atendente.get_full_name() if lead.atendente else '',
            lead.get_status_display(),
            lead.data_criacao.strftime('%d/%m/%Y %H:%M'),
            lead.observacao_interna or ''
        ]

        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border

    # Ajustar larguras das colunas
    column_widths = {
        'A': 25,  # Nome
        'B': 15,  # Telefone
        'C': 25,  # Email
        'D': 20,  # Curso
        'E': 12,  # Origem
        'F': 12,  # Prioridade
        'G': 12,  # Probabilidade
        'H': 15,  # Valor Curso
        'I': 20,  # Atendente
        'J': 12,  # Status
        'K': 18,  # Data Criação
        'L': 30,  # Observação Interna
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Congelar a primeira linha
    ws.freeze_panes = 'A2'

    # Salvar em buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Criar resposta HTTP
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="leads.xlsx"'

    return response


@login_required
def export_leads_pdf(request):
    # Função auxiliar para criar parágrafos que se ajustam às células
    def create_cell_paragraph(text, max_length=None, font_size=8):
        if not text or text == '-':
            return text or '-'

        # Limitar comprimento se especificado
        if max_length and len(text) > max_length:
            text = text[:max_length] + '...'

        # Criar estilo para célula
        cell_style = ParagraphStyle(
            'CellStyle',
            fontSize=font_size,
            leading=font_size + 2,  # Espaçamento entre linhas
            alignment=0,  # Esquerda
        )

        return Paragraph(text, cell_style)

    # Aplicar os mesmos filtros da listagem
    leads = Lead.objects.all()

    # Filtrar por atendente se for ATENDENTE
    if request.user.groups.filter(name='ATENDENTE').exists():
        leads = leads.filter(atendente=request.user)

    # Aplicar filtros da requisição (mesmo código do CSV)
    query = request.GET.get("q", "")
    status_filter = request.GET.get("status", "")
    origem_filter = request.GET.get("origem", "")
    atendente_filter = request.GET.get("atendente", "")
    prioridade_filter = request.GET.get("prioridade", "")
    prob_min = request.GET.get("prob_min", "")
    prob_max = request.GET.get("prob_max", "")
    data_inicial = request.GET.get("data_inicial", "")
    data_final = request.GET.get("data_final", "")

    if query:
        leads = leads.filter(Q(nome__icontains=query) | Q(email__icontains=query) | Q(curso_interesse__icontains=query))

    if status_filter:
        leads = leads.filter(status=status_filter)

    if origem_filter:
        leads = leads.filter(origem=origem_filter)

    if atendente_filter:
        leads = leads.filter(atendente_id=atendente_filter)

    if prioridade_filter:
        leads = leads.filter(prioridade=prioridade_filter)

    if prob_min:
        leads = leads.filter(probabilidade_fechamento__gte=prob_min)

    if prob_max:
        leads = leads.filter(probabilidade_fechamento__lte=prob_max)

    if data_inicial:
        leads = leads.filter(data_criacao__date__gte=data_inicial)

    if data_final:
        leads = leads.filter(data_criacao__date__lte=data_final)

    # Criar PDF com orientação paisagem para mais espaço
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []

    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Centralizado
    )

    # Título
    title = Paragraph("Relatório de Leads", title_style)
    elements.append(title)

    # Informações do filtro
    filter_info = []
    if query:
        filter_info.append(f"Busca: {query}")
    if status_filter:
        filter_info.append(f"Status: {dict(Lead.STATUS_CHOICES).get(status_filter, status_filter)}")
    if origem_filter:
        filter_info.append(f"Origem: {dict(Lead.ORIGEM_CHOICES).get(origem_filter, origem_filter)}")
    if atendente_filter:
        try:
            user = User.objects.get(id=atendente_filter)
            filter_info.append(f"Atendente: {user.get_full_name() or user.username}")
        except:
            filter_info.append(f"Atendente ID: {atendente_filter}")

    if filter_info:
        filter_text = "Filtros aplicados: " + ", ".join(filter_info)
        filter_paragraph = Paragraph(filter_text, styles['Normal'])
        elements.append(filter_paragraph)
        elements.append(Spacer(1, 12))

    # Estilo para cabeçalhos da tabela
    header_style = ParagraphStyle(
        'HeaderStyle',
        fontSize=9,
        fontName='Helvetica-Bold',
        alignment=1,  # Centralizado
        textColor=colors.whitesmoke
    )

    # Dados da tabela com parágrafos para quebra automática de linhas
    data = [
        [
            Paragraph('Nome', header_style),
            Paragraph('Telefone', header_style),
            Paragraph('Email', header_style),
            Paragraph('Curso', header_style),
            Paragraph('Origem', header_style),
            Paragraph('Prioridade', header_style),
            Paragraph('Prob.', header_style),
            Paragraph('Valor', header_style),
            Paragraph('Atendente', header_style),
            Paragraph('Status', header_style),
            Paragraph('Data', header_style)
        ]
    ]

    for lead in leads:
        # Usar parágrafos para permitir quebra automática de linhas
        data.append([
            create_cell_paragraph(lead.nome, max_length=30),
            create_cell_paragraph(lead.telefone),
            create_cell_paragraph(lead.email or '-', max_length=25),
            create_cell_paragraph(lead.curso_interesse or '-', max_length=25),
            create_cell_paragraph(lead.get_origem_display()),
            create_cell_paragraph(lead.get_prioridade_display()),
            create_cell_paragraph(f"{lead.probabilidade_fechamento}%"),
            create_cell_paragraph(f"R$ {lead.valor_curso:,}" if lead.valor_curso else '-'),
            create_cell_paragraph(lead.atendente.get_full_name() if lead.atendente else '-', max_length=20),
            create_cell_paragraph(lead.get_status_display()),
            create_cell_paragraph(lead.data_criacao.strftime('%d/%m/%Y'))
        ])

    # Criar tabela com larguras de coluna ajustadas
    col_widths = [1.4*inch, 0.9*inch, 1.4*inch, 1.2*inch, 0.8*inch, 0.8*inch, 0.6*inch, 0.9*inch, 1.2*inch, 0.8*inch, 0.8*inch]
    table = Table(data, colWidths=col_widths)

    # Estilo da tabela melhorado
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 1), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 2),
    ]))

    elements.append(table)

    # Adicionar total de leads
    total_style = ParagraphStyle(
        'TotalStyle',
        parent=styles['Normal'],
        fontSize=10,
        spaceBefore=20,
        alignment=2  # Direita
    )
    total_text = f"Total de leads: {leads.count()}"
    total_paragraph = Paragraph(total_text, total_style)
    elements.append(total_paragraph)

    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="leads.pdf"'

    return response


@login_required
def lead_create(request):
    if request.method == "POST":
        form = LeadForm(request.POST, user=request.user)
        if form.is_valid():
            lead = form.save()
            # Criar log de atividade
            ActivityLog.objects.create(
                lead=lead,
                user=request.user,
                action='created',
                description=f'Lead criado por {request.user.get_full_name() or request.user.username}'
            )
            messages.success(request, f'Lead "{lead.nome}" criado com sucesso!')
            return redirect("leads:lead_list")
    else:
        form = LeadForm(user=request.user)
    return render(request, "leads/form.html", {"form": form, "title": "Criar Lead"})


@login_required
def lead_detail(request, pk):
    lead = get_object_or_404(Lead, pk=pk)
    # Verificar se ATENDENTE pode ver apenas leads próprios
    if request.user.groups.filter(name='ATENDENTE').exists() and lead.atendente != request.user:
        messages.error(request, 'Você não tem permissão para visualizar este lead.')
        return redirect('leads:lead_list')
    return render(request, "leads/detail.html", {"lead": lead})


@login_required
def lead_edit(request, pk):
    lead = get_object_or_404(Lead, pk=pk)
    # Verificar se ATENDENTE pode editar apenas leads próprios
    if request.user.groups.filter(name='ATENDENTE').exists() and lead.atendente != request.user:
        messages.error(request, 'Você não tem permissão para editar este lead.')
        return redirect('leads:lead_list')
    if request.method == "POST":
        old_status = lead.status
        old_atendente = lead.atendente
        form = LeadForm(request.POST, instance=lead, user=request.user)
        if form.is_valid():
            lead = form.save()

            # Criar logs de atividade
            if old_status != lead.status:
                ActivityLog.objects.create(
                    lead=lead,
                    user=request.user,
                    action='status_changed',
                    old_value=old_status,
                    new_value=lead.status,
                    description=f'Status alterado de "{dict(Lead.STATUS_CHOICES)[old_status]}" para "{dict(Lead.STATUS_CHOICES)[lead.status]}"'
                )

            if old_atendente != lead.atendente:
                ActivityLog.objects.create(
                    lead=lead,
                    user=request.user,
                    action='assigned',
                    old_value=str(old_atendente) if old_atendente else None,
                    new_value=str(lead.atendente) if lead.atendente else None,
                    description=f'Atendente alterado para {lead.atendente.get_full_name() if lead.atendente else "Nenhum"}'
                )

            if old_status == lead.status and old_atendente == lead.atendente:
                ActivityLog.objects.create(
                    lead=lead,
                    user=request.user,
                    action='updated',
                    description=f'Lead atualizado por {request.user.get_full_name() or request.user.username}'
                )

            messages.success(request, f'Lead "{lead.nome}" atualizado com sucesso!')
            return redirect("leads:lead_detail", pk=pk)
    else:
        form = LeadForm(instance=lead, user=request.user)
    return render(request, "leads/form.html", {"form": form, "title": "Editar Lead"})


@login_required
def lead_delete(request, pk):
    lead = get_object_or_404(Lead, pk=pk)
    # Verificar se ATENDENTE pode deletar apenas leads próprios
    if request.user.groups.filter(name='ATENDENTE').exists() and lead.atendente != request.user:
        messages.error(request, 'Você não tem permissão para deletar este lead.')
        return redirect('leads:lead_list')
    if request.method == "POST":
        # Criar log antes de deletar
        ActivityLog.objects.create(
            lead=lead,
            user=request.user,
            action='deleted',
            description=f'Lead deletado por {request.user.get_full_name() or request.user.username}'
        )
        lead.delete()
        messages.success(request, f'Lead "{lead.nome}" deletado com sucesso!')
        return redirect("leads:lead_list")
    return render(request, "leads/delete.html", {"lead": lead})
